"""Document text extraction + image preparation for vision models."""

import base64
import logging
import os
from io import BytesIO

from ..config import settings

log = logging.getLogger(__name__)

IMAGE_MIMES = {"image/png", "image/jpeg", "image/webp", "image/gif"}
TEXT_MIMES = {"text/plain", "text/csv", "text/markdown", "application/json"}
PDF_MIME = "application/pdf"
DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MIME_BY_EXT = {
    ".pdf": PDF_MIME,
    ".docx": DOCX_MIME,
    ".xlsx": XLSX_MIME,
    ".csv": "text/csv",
    ".txt": "text/plain",
    ".md": "text/markdown",
    ".json": "application/json",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".webp": "image/webp",
    ".gif": "image/gif",
}


def detect_mime(filename: str, content_type: str | None) -> str:
    ct = (content_type or "").split(";")[0].strip().lower()
    if ct and ct != "application/octet-stream" and not ct.startswith("text/xml"):
        return ct
    return MIME_BY_EXT.get(os.path.splitext(filename.lower())[1], ct)


def allowed_mime(mime: str) -> bool:
    return mime in IMAGE_MIMES or mime in TEXT_MIMES or mime in {PDF_MIME, DOCX_MIME, XLSX_MIME}


def _truncate(text: str, limit: int) -> str:
    if len(text) <= limit:
        return text
    return text[:limit] + f"\n\n…[truncated {len(text) - limit} chars]"


def extract_text(path: str, mime: str) -> str | None:
    """Synchronous (CPU) — call via asyncio.to_thread."""
    try:
        if mime == PDF_MIME:
            from pypdf import PdfReader

            reader = PdfReader(path)
            text = "\n\n".join((page.extract_text() or "") for page in reader.pages)
        elif mime == DOCX_MIME:
            from docx import Document

            text = "\n".join(p.text for p in Document(path).paragraphs)
        elif mime == XLSX_MIME:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            parts: list[str] = []
            for ws in wb.worksheets:
                parts.append(f"# Sheet: {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    parts.append("\t".join("" if c is None else str(c) for c in row))
            text = "\n".join(parts)
        elif mime in TEXT_MIMES:
            with open(path, "r", encoding="utf-8", errors="ignore") as fh:
                text = fh.read()
        else:
            return None
        return _truncate(text.strip(), settings.MAX_FILE_CHARS) if text else None
    except Exception as e:
        log.warning("text extraction failed for %s: %s", path, e)
        return None


def image_data_url(path: str, mime: str, max_dim: int = 1600) -> str:
    """Downscale and base64-encode an image for the vision API (caps token cost)."""
    from PIL import Image

    im = Image.open(path)
    im.thumbnail((max_dim, max_dim))
    use_png = mime == "image/png" and im.mode in ("RGBA", "P", "LA")
    if not use_png and im.mode != "RGB":
        im = im.convert("RGB")
    buf = BytesIO()
    im.save(buf, "PNG" if use_png else "JPEG", quality=88)
    b64 = base64.b64encode(buf.getvalue()).decode()
    return f"data:image/{'png' if use_png else 'jpeg'};base64,{b64}"
